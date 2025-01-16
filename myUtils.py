#!/usr/bin/env python
# coding=utf-8
import copy
import datetime
import json
import os
import queue
import re
import socket
import traceback

import openpyxl as oxl
import requests
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem
# 全局变量区域
from openpyxl.styles import Border, Side, Font, PatternFill

borderNumDic = {-1: None, 0: "thin"}
# 降低SECLEVEL级别，防止SSL异常
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'DEFAULT:@SECLEVEL=1'


def createAnalysisJsonResultItem():
    reDict = {}

    reDict["level"] = -1
    reDict["value"] = ""
    reDict["nextItems"] = {}

    return reDict

def getSplitChar():
    return "☯"

# 处理json对象，返回一个用字典存储的树状结构
# 返回结果字典，列表中存储一个树状结构，每个节点代表一个json item
# 节点key的格式为：<父节点类型(list、dict、root)>_<父节点获取该子节点的关键词，dict为key值，list为index>_<该节点的类型，字典为dict，列表为list，其他值为该值的类型：str、int、float等>，节点的值为一个字典结构，由createAnalysisJsonResultItem函数定义
# 例：
# {"a":{"b":"c","d":[1,2,"3"]},"e":1.1}
# {'root_0_dict': {'level': 0, 'value': '2', 'nextItems': {'dict_a_dict': {'level': 1, 'value': '2', 'nextItems': {'dict_b_str': {'level': 2, 'value': 'c', 'nextItems': {}}, 'dict_d_list': {'level': 2, 'value': '3', 'nextItems': {'list_0_int': {'level': 3, 'value': '1', 'nextItems': {}}, 'list_1_int': {'level': 3, 'value': '2', 'nextItems': {}}, 'list_2_str': {'level': 3, 'value': '3', 'nextItems': {}}}}}}, 'dict_e_float': {'level': 1, 'value': '1.1', 'nextItems': {}}}}}
def analysisJsonObjToDict(jsonObj):
    reDict = {}

    connectStr = getSplitChar()

    anaQueue = queue.Queue()

    anaQueue.put({"obj":jsonObj,"resultDict":reDict,"level":0,"keyStr":f"root{connectStr}0"})

    while not anaQueue.empty():
        tmpObjDict = anaQueue.get()
        tmpJsonObj = tmpObjDict["obj"]
        tmpResultDict = tmpObjDict["resultDict"]
        tmpLevel = tmpObjDict["level"]
        tmpKeyStr = tmpObjDict["keyStr"]

        if type(tmpJsonObj) == dict:
            tmpResultKey = f"{tmpKeyStr}{connectStr}dict"
            tmpResultItem = createAnalysisJsonResultItem()
            tmpResultItem["level"] = tmpLevel
            tmpResultItem["value"] = str(len(tmpJsonObj.keys()))
            tmpResultDict[tmpResultKey] = tmpResultItem
            for tmpKey,tmpVal in tmpJsonObj.items():
                anaQueue.put({"obj": copy.deepcopy(tmpVal), "resultDict": tmpResultDict[tmpResultKey]["nextItems"], "level": tmpLevel+1,"keyStr":f"dict{connectStr}{tmpKey}"})
        elif type(tmpJsonObj) == list:
            tmpResultKey = f"{tmpKeyStr}{connectStr}list"
            tmpResultItem = createAnalysisJsonResultItem()
            tmpResultItem["level"] = tmpLevel
            tmpResultItem["value"] = str(len(tmpJsonObj))
            tmpResultDict[tmpResultKey] = tmpResultItem
            for tmpIndex,tmpVal in enumerate(tmpJsonObj):
                anaQueue.put(
                    {"obj": copy.deepcopy(tmpVal), "resultDict": tmpResultDict[tmpResultKey]["nextItems"], "level": tmpLevel + 1,
                     "keyStr": f"list{connectStr}{tmpIndex}"})
        else:
            tmpResultKey = f"{tmpKeyStr}{connectStr}{type(tmpJsonObj).__name__}"
            tmpResultItem = createAnalysisJsonResultItem()
            tmpResultItem["level"] = tmpLevel
            tmpResultItem["value"] = str(tmpJsonObj)
            tmpResultDict[tmpResultKey] = tmpResultItem

    return reDict


# 判断传入的对象类型，是dict返回1，是list返回2，否则返回0
def checkObjType(obj):
    retStatus = 0
    if type(obj) == dict:
        retStatus = 1
    elif type(obj) == list:
        retStatus = 2
    else:
        retStatus = 0
    return retStatus


# 根据传入的contentDic写入配置文件，若配置文件不存在会创建
# contentDic格式为：{"配置名":"配置值"}，
# 对list和dict格式的值会使用json.dumps进行转换
def writeToConfFile(filePath, contentDic):
    with open(filePath, "w+", encoding="utf-8") as fr:
        for key, value in contentDic.items():
            if isinstance(value, list) or isinstance(value, dict):
                value = json.dumps(value)
            content = "{0}={1}".format(key, value)
            fr.write(content + "\n")


# 读取配置文件，生成一个配置字典，
# 字典结构为：{"配置名":"配置值"}，
# 文件结构为：配置名=配置值，对list和dict格式的值会使用json.loads进行转换
# 配置字典的最后一项固定为"confHeader":配置名列表
def readConfFile(filePath):
    confDic = {}
    headerList = []
    with open(filePath, "r", encoding="utf-8") as fr:
        fileLines = fr.readlines()
    fileLines = [a.replace("\r\n", "\n").replace("\n", "") for a in fileLines if a != ""]
    for fileLine in fileLines:
        fileLine = fileLine.replace("\r\n", "\n").replace("\n", "")
        tempList = fileLine.split("=")
        key = tempList[0].strip()
        value = "=".join(tempList[1:]).strip()
        try:
            value = json.loads(value)
        except:
            pass
        confDic[key] = value
        headerList.append(key)
    confDic["confHeader"] = headerList

    return confDic


# 判断该字符串是否是IP，返回一个布尔值
def ifIp(matchStr):
    reFlag = True
    ipReg = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    if re.match(ipReg, matchStr):
        reFlag = True
    else:
        reFlag = False
    return reFlag


# 测试指定IP的指定端口是否能联通
def connectIpPort(ip, port):
    ifConnected = False
    socketObj = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = socketObj.connect_ex((ip, port))
    if result == 0:
        ifConnected = True
    else:
        ifConnected = False
    return ifConnected


# 获得精确到秒的当前时间
def getNowSeconed():
    formatStr = "%Y-%m-%d %H:%M:%S"
    nowDate = datetime.datetime.now()
    nowDateStr = nowDate.strftime(formatStr)
    return nowDateStr


# 访问URL,type表示请求类型，0为GET，1为POST，2为PUT
# 返回值类型如下：
# {
# "url":传入URL,
# "resultStr":访问结果字符串,
# "checkFlag":标志是否访问成功的布尔类型变量,
# "title":访问成功时的页面标题,
# "pageContent":访问成功时的页面源码，
# "status":访问的响应码，
# "requestSeconed":访问耗时，单位为秒
# }
def requestsUrl(url, cookie={}, header={}, data={}, files=None, type=0, reqTimeout=10, readTimeout=120,
                allow_redirects=False, session=None, proxies=None):
    if session is None:
        session = requests.session()
    else:
        pass
    resDic = {}
    url = url.strip()
    url = url.strip()

    resultStr = ""
    checkedFlag = False
    status = ""
    title = ""
    reContent = ""
    totalSeconed = 0
    timeout = (reqTimeout, readTimeout)
    header = header if header != {} else {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"
    }
    try:
        if type == 0:
            response = session.get(url, headers=header, verify=False, cookies=cookie, timeout=timeout,
                                   allow_redirects=allow_redirects, proxies=proxies)
        elif type == 1:
            response = session.post(url, headers=header, verify=False, cookies=cookie, data=data, files=files,
                                    timeout=timeout, allow_redirects=allow_redirects, proxies=proxies)
        elif type == 2:
            response = session.put(url, headers=header, verify=False, cookies=cookie, data=data, files=files,
                                   timeout=timeout, allow_redirects=allow_redirects, proxies=proxies)
        else:
            pass
        status = response.status_code
        totalSeconed = response.elapsed.total_seconds()
        if str(status)[0] == "2" or str(status)[0] == "3":
            # 获得页面编码
            pageEncoding = response.apparent_encoding
            # 设置页面编码
            response.encoding = pageEncoding
            # 获得页面内容
            reContent = response.text
            reTitleList = re.findall(r"<title.*?>(.+?)</title>", reContent)
            title = "成功访问，但无法获得标题" if len(reTitleList) == 0 else reTitleList[0]
            resultStr = "验证成功，标题为：{0}".format(title)
            checkedFlag = True
        else:
            resultStr = "验证失败，状态码为{0}".format(status)
            checkedFlag = False
    except Exception as e:
        resultStr = traceback.format_exc()
        checkedFlag = False

    # 构建返回结果
    resDic["url"] = url
    resDic["resultStr"] = resultStr
    resDic["checkFlag"] = checkedFlag
    resDic["title"] = title
    resDic["status"] = status
    resDic["pageContent"] = reContent
    resDic["requestSeconed"] = totalSeconed
    return resDic


# 传入一个请求包格式字符串，解析并返回一个格式化的请求包字典，用于进行requests请求
# 可以设置header字段中不读取的字段名，默认不读取Content-Length
# 返回请求包字典格式为：
# {
#   "requestType":返回请求方式，0为GET，1为POST，2为PUT，其他返回-1
#   "host":请求的域名（不包含路径部分）,
#   "uri":请求的路径（不包含域名部分）,
#   "header":请求包中的header部分，是一个字典,
#   "data":请求包中的数据行，若不存在数据行则返回一个空字符串,
#   "cookie":请求包中的cookie,是字典类型变量
# }
def analysisPacketFromTxt(packetTxt, notReadKeys=["Content-Length"]):
    reDic = {}
    requestTypeDic = {"get": 0, "post": 1, "put": 2}
    # 解析报文内容
    packetTxt = packetTxt.replace("\r\n", "\n")
    packetContents = packetTxt.split("\n")
    # 判断请求方式和请求地址
    requestTypeTxt = packetContents[0].split(" ")[0].lower()
    if requestTypeTxt in requestTypeDic.keys():
        requestType = requestTypeDic[requestTypeTxt]
    else:
        requestType = -1
    uri = packetContents[0].split(" ")[1]
    host = ""
    cookieStr = ""

    # 生成header和data行的数据
    header = {}
    dataLineContent = ""
    for index in range(1, len(packetContents)):
        nowLine = packetContents[index]
        nowLine = nowLine.replace("\n", "")
        # 发现当前行为空行，且不是数据包的最后一行，判定下面的内容中存在数据行,执行完后结束循环
        if nowLine == "" and index != len(packetContents) - 1:
            for tmpIndex in range(1, len(packetContents) - index):
                tmpLine = packetContents[index + 1].replace("\n", "")
                if tmpLine == "":
                    continue
                else:
                    dataLineContent = tmpLine
                    break
            break
        tempArr = nowLine.split(":")
        nowKey = tempArr[0]
        nowValue = ":".join(tempArr[1:]).strip()

        if nowKey.lower() == "cookie":
            cookieStr = nowValue
            continue

        if nowKey.lower() == "host":
            host = nowValue
            continue

        if nowKey not in notReadKeys:
            header.update({nowKey: nowValue})

    # 处理cookie字符串
    cookie = {}
    if cookieStr!="":
        tmpList = cookieStr.split(";")
        for tmpStr in tmpList:
            tmpCookieItem = tmpStr.split("=")
            nowKey = tmpCookieItem[0]
            nowValue = "=".join(tmpCookieItem[1:])
            cookie[nowKey] = nowValue

    # 生成返回结果字典
    reDic["requestType"] = requestType
    reDic["host"] = host
    reDic["uri"] = uri
    reDic["header"] = header
    reDic["data"] = dataLineContent.encode("utf-8")
    reDic["cookie"] = cookie
    return reDic


# 创建一个QTableWidgetItem对象，传入文本参数text，以及verAlign和horAlign两个整形参数控制文本位置
# verAlign的值含义如下：-1 靠左、0 居中、1 靠右
# horAlign的值含义如下：-1 靠下、0 居中、1 靠上
def createTableItem(text, verAlign=0, horAlign=0):
    tmpItem = QTableWidgetItem(str(text))
    verAlignDic = {-1: Qt.AlignLeft, 0: Qt.AlignVCenter, 1: Qt.AlignRight}
    horAlignDic = {-1: Qt.AlignBottom, 0: Qt.AlignHCenter, 1: Qt.AlignTop}
    tmpItem.setTextAlignment(verAlignDic[verAlign] | horAlignDic[horAlign])
    return tmpItem


# 传入一个QTableWidget对象，清空这个表格
def clearTalbe(tableObject):
    while tableObject.rowCount() != 0:
        tableObject.removeRow(tableObject.rowCount() - 1)


# 用于进行动态坐标系加1，传入一个当前坐标值的列表，以及一个最大坐标系长度的列表，按从后到前的顺序进行计算（例子：多重for嵌套）
# 返回一个布尔类型参数ifAddFlag和一个进行坐标系加法后的坐标值列表，当ifAddFlag为True时表示进行了坐标系加法，否则表示所有坐标已遍历完
def coordinatePlus(nowCoordinate, maxCoordinate):
    ifAddFlag = False
    nowAddIndex = -1
    while not ifAddFlag:
        if len(nowCoordinate) + nowAddIndex == -1:
            ifAddFlag = False
            break
        else:
            nowCoordinateItem = nowCoordinate[nowAddIndex]
            nowMaxCoordinateItem = maxCoordinate[nowAddIndex]
            if nowCoordinateItem == nowMaxCoordinateItem:
                tmpAddIndex = nowAddIndex
                while tmpAddIndex != 0:
                    nowCoordinate[tmpAddIndex] = 0
                    tmpAddIndex = tmpAddIndex + 1
                nowAddIndex = nowAddIndex - 1
            else:
                nowCoordinate[nowAddIndex] = nowCoordinate[nowAddIndex] + 1
                ifAddFlag = True
    return ifAddFlag, nowCoordinate


# 获得excell的常用样式
def getExcellStyleDic():
    styleDic = {}

    # 单线边框
    thinBorder = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

    # 文字居中
    alignStyle = oxl.styles.Alignment(horizontal='center', vertical='center')
    leftStyle = oxl.styles.Alignment(horizontal='left', vertical='center')
    rightStyle = oxl.styles.Alignment(horizontal='right', vertical='center')

    # 加粗字体
    boldFont = Font(bold=True)
    hyperLinkFont = Font(color='0000FF')
    underLineFont = Font(underline='single')

    styleDic["thin"] = thinBorder
    styleDic["align"] = alignStyle
    styleDic["bold"] = boldFont
    styleDic["left"] = leftStyle
    styleDic["right"] = rightStyle
    styleDic["link"] = hyperLinkFont
    styleDic["underLine"] = underLineFont
    return styleDic


# 写入一个标准的excell表头（居中，单线框，加粗）
def writeExcellHead(ws, headArr):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 写入表头
    for index, head in enumerate(headArr):
        ws.cell(row=1, column=index + 1).value = head
        ws.cell(row=1, column=index + 1).border = styleDic["thin"]
        ws.cell(row=1, column=index + 1).alignment = styleDic["align"]
        ws.cell(row=1, column=index + 1).font = styleDic["bold"]
    return ws


# 写入一个内容单元格
# borderNum表示该单元格的边框对象，其值可查询全局变量styleDic
# ifAlign是一个boolean对象，True表示居中
# hyperLink表示该单元格指向的链接，默认为None，表示不指向任何链接
# fgColor表示该单元格的背景颜色，为一个RGB16进制字符串，默认为“FFFFFF”（白色）
# otherAlign表示当ifAlign为False时指定的其他对齐方式，是一个数字型变量，默认为None，当其为0时表示左对齐，1为右对齐
def writeExcellCell(ws, row, column, value, borderNum, ifAlign, hyperLink=None, fgColor="FFFFFF", otherAlign=None):
    value = str(value)
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 获得指定单元格
    aimCell = ws.cell(row=row, column=column)
    # 设置值
    aimCell.value = value
    # 设置边框
    styleObjKey = borderNumDic[borderNum]
    if not styleObjKey:
        pass;
    else:
        styleObj = styleDic[styleObjKey]
        aimCell.border = styleObj
    # 设置居中
    if ifAlign:
        aimCell.alignment = styleDic["align"]
    elif otherAlign is not None:
        otherAlign = int(otherAlign)
        if otherAlign == 0:
            aimCell.alignment = styleDic["left"]
        else:
            aimCell.alignment = styleDic["right"]
    else:
        pass

    # 设置超链接
    if hyperLink:
        # 写入超链接
        aimCell.hyperlink = hyperLink
        # 设置当前单元格字体颜色为深蓝色，并添加下划线
        aimCell.font = styleDic["link"]
    else:
        pass

    # 设置填充颜色
    fill = PatternFill("solid", fgColor=fgColor)
    aimCell.fill = fill

    return ws


# 写入一个空格单元格，防止上一列文本超出
def writeExcellSpaceCell(ws, row, column):
    # 设置值
    ws.cell(row=row, column=column).value = " "

    return ws


# 设置excell的列宽
def setExcellColWidth(ws, colWidthArr):
    for colWidindex in range(len(colWidthArr)):
        ws.column_dimensions[chr(ord("A") + colWidindex)].width = colWidthArr[colWidindex]

    return ws


# 保存excell文件
def saveExcell(wb, saveName):
    savePath = ""
    # 处理传入的文件名
    saveName = saveName.split(".")[0] + ".xlsx"
    savePath = "{0}\\{1}".format(os.getcwd(), saveName)

    # 检测当前目录下是否有该文件，如果有则清除以前保存文件
    if os.path.exists(savePath):
        deleteFile(savePath)
    wb.save(savePath)
    return True


# 删除指定路径的文件,传入一个绝对路径,返回一个布尔变量以及一个字符串变量，
# 布尔变量为True表示是否删除成功,若为False则字符串变量中写入错误信息
def deleteFile(filePath):
    deleteFlag = True
    reStr = ""
    if os.path.exists(filePath):
        try:
            os.remove(filePath)
        except Exception as ex:
            reStr = "删除失败，失败信息为：{0}".format(ex)
            deleteFlag = False
    else:
        reStr = "未找到指定路径的文件"
        deleteFlag = False
    return deleteFlag, reStr


# 将传入的字符串修改为符合windows文件名规范的字符串
def updateFileNameStr(oriStr):
    resultStr = oriStr
    # 替换换行符
    resultStr = resultStr.replace("\r\n", "\n").replace("\n", "")
    # 将违法字符替换为下划线
    notAllowCharList = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|", "-"]
    for nowNotAllowChar in notAllowCharList:
        resultStr = resultStr.replace(nowNotAllowChar, "_")
    return resultStr