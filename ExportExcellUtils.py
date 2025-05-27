#!/usr/bin/env python
# coding=utf-8
import os
import re
import traceback
import datetime

import openpyxl as oxl
from openpyxl.styles import Border, Side, Font, PatternFill

import warnings

warnings.filterwarnings("ignore")


# 使用方法：
# exportExcellUtils = ExportExcellUtils(saveCount=10000)   # 实例化一个导出工具类
# nowExcellFileObj = exportExcellUtils.addFile(fileName)  # 添加一个excell文件，传入文件名
# nowExcellSheetObj = nowExcellFileObj.getFinalSheet()   # 获取最后一个sheet，第一个sheet会自动创建
# nowExcellSheetObj.sheetName = fileName   # 设置sheet名
# nowExcellSheetObj.setHeaderList(exportExcellUtils.transformListToHeaderList(header))  # 设置sheet的表头，表头不用包含序号，序号会自动添加
# nowExcellSheetObj.addRows([exportExcellUtils.transformListToCellList(tmpResults) for tmpResults in solvedLogResults(resultLine)])  # 传入需要写入的数据列表
# exportExcellUtils.exportExcell(fileIndex)   # 保存excell文件

# excell 单元格类
class ExcellCell:
    def __init__(self, value="", border="thin", fontAlignment="align", fontStyle="none",
                 fgColor="white", hyperLink=None):
        self.value = str(value)
        self.border = border
        self.fontAlignment = fontAlignment
        self.fontStyle = fontStyle
        self.fgColor = fgColor
        self.hyperLink = hyperLink
        self.cellWidth = self.calcCellWidth()

    def calcCellWidth(self):
        cellWidth = 0
        for tmpChar in self.value:
            if tmpChar.isascii():
                cellWidth += 1
            else:
                cellWidth += 3
        return cellWidth + 4

    def getCellWidth(self):
        return self.cellWidth

    def setValue(self, value):
        self.value = value
        self.cellWidth = self.calcCellWidth()

    def setCellAsHeader(self):
        self.border = "thin"
        self.fontAlignment = "align"
        self.fontStyle = "bold"


# excell 子表类
class ExcellSheet:
    def __init__(self, sheetName="sheet1", headerList=[""]):
        self.sheetName = sheetName
        self.headerList = []
        self.setHeaderList(headerList)
        self.dataList = []

    def setHeaderList(self, headerList):
        tmpIndexCellItem = ExcellCell(value="序号")
        tmpIndexCellItem.setCellAsHeader()
        self.headerList = [tmpIndexCellItem]
        self.headerList += headerList

    def addRow(self, items):
        tmpRowList = [ExcellCell(value=len(self.dataList) + 1)]
        for item in items:
            tmpRowList.append(item)
        self.dataList.append(tmpRowList)

    def addRows(self, rows):
        for tmpRowList in rows:
            self.addRow(tmpRowList)

    def getRows(self):
        return self.dataList

    def getSheetName(self):
        return self.sheetName

    def getHeaderList(self):
        return self.headerList


# excell 文件类
class ExcellFile:
    def __init__(self, fileName="导出文件", fileSuffix="xlsx", sheetName=""):
        saveName = fileName + "-" + self.getNowSeconed() + "." + fileSuffix
        self.saveName = self.updateFileNameStr(saveName)
        self.sheets = []
        self.addSheet(sheetName)

    # 获得精确到秒的当前时间
    def getNowSeconed(self):
        formatStr = "%Y-%m-%d %H:%M:%S"
        nowDate = datetime.datetime.now()
        nowDateStr = nowDate.strftime(formatStr)
        return nowDateStr

    # 将传入的字符串修改为符合windows文件名规范的字符串
    def updateFileNameStr(self, oriStr):
        resultStr = oriStr
        # 替换换行符
        resultStr = resultStr.replace("\r\n", "\n").replace("\n", "")
        # 将违法字符替换为下划线
        notAllowCharList = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|", "-"]
        for nowNotAllowChar in notAllowCharList:
            resultStr = resultStr.replace(nowNotAllowChar, "_")
        return resultStr

    # 创建一个sheet并返回该对象
    def addSheet(self, sheetName, headerList=[""], rows=[[]]):
        if sheetName == "":
            sheetName = "sheet{}".format(len(self.sheets) + 1)

        tmpSheet = ExcellSheet(sheetName)
        self.sheets.append(tmpSheet)
        return tmpSheet

    # 返回当前文件最后一个sheet对象
    def getFinalSheet(self):
        return self.sheets[len(self.sheets) - 1]

    # 返回当前文件保存文件名（包含后缀）
    def getSaveName(self):
        return self.saveName

    # 返回当前的sheet列表
    def getSheets(self):
        return self.sheets

    # 通过子表名获取对应子表对象
    def getSheetByName(self, aimSheetName):
        aimSheet = None
        nowSheets = self.getSheets()
        for nowSheet in nowSheets:
            nowSheetName = nowSheet.getSheetName()
            if aimSheetName == nowSheetName:
                aimSheet = nowSheet
                break
            else:
                continue
        return aimSheet

    # 通过子表索引获取对应子表对象，索引从0开始
    def getSheetByIndex(self, sheetIndex):
        aimSheet = None
        nowSheets = self.getSheets()
        if sheetIndex < len(nowSheets):
            aimSheet = nowSheets[sheetIndex]
        else:
            pass
        return aimSheet


# 用于导出数据为excell的工具类
class ExportExcellUtils:
    def __init__(self, saveCount=1000):
        self.saveCount = saveCount
        self.fileList = []
        self.styleDict = {}

        self.initExcellStyleDic()

    # 初始化excell的常用样式
    def initExcellStyleDic(self):
        styleDict = {"borders": None, "fontAlignments": None, "fontStyles": None, "colorCodes": None,
                     "cellWidths": None}
        borders = {}
        fontAlignments = {}
        fontStyles = {}
        colorCodes = {}
        cellWidths = {}

        # 颜色
        colorCodes["white"] = "FFFFFF"
        colorCodes["black"] = "000000"
        colorCodes["blue"] = "0000FF"
        colorCodes["green"] = "008000"
        colorCodes["gray"] = "808080"
        colorCodes["yellow"] = "FFFF00"
        colorCodes["red"] = "FF0000"

        # 单线边框
        thinBorder = Border(left=Side(border_style='thin', color=colorCodes["black"]),
                            right=Side(border_style='thin', color=colorCodes["black"]),
                            top=Side(border_style='thin', color=colorCodes["black"]),
                            bottom=Side(border_style='thin', color=colorCodes["black"]))
        noneBorder = Border()
        borders["thin"] = thinBorder
        borders["none"] = noneBorder

        # 文字居中
        alignStyle = oxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        leftStyle = oxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        rightStyle = oxl.styles.Alignment(horizontal='right', vertical='center', wrap_text=True)
        fontAlignments["align"] = alignStyle
        fontAlignments["left"] = leftStyle
        fontAlignments["right"] = rightStyle

        # 字体
        boldFont = Font(bold=True)  # 加粗
        hyperLinkFont = Font(colorCodes["blue"])  # 超链接用,字体颜色为深蓝色
        underLineFont = Font(underline='single')  # 下划线
        noneFontStyle = Font()
        fontStyles["bold"] = boldFont
        fontStyles["hyper"] = hyperLinkFont
        fontStyles["underLine"] = underLineFont
        fontStyles["none"] = noneFontStyle

        # 单元格宽度
        cellWidths["default"] = 10
        cellWidths["index"] = 7
        cellWidths["middle"] = 30
        cellWidths["long"] = 70
        cellWidths["max"] = 70  # 自动调节的最大列宽

        styleDict["borders"] = borders
        styleDict["fontAlignments"] = fontAlignments
        styleDict["fontStyles"] = fontStyles
        styleDict["colorCodes"] = colorCodes
        styleDict["cellWidths"] = cellWidths

        self.styleDict = styleDict

    # 返回excell的常用样式字典
    def getStyleDict(self):
        return self.styleDict

    # 传入一个值的列表，将其转换成excell表头列表
    def transformListToHeaderList(self, valueList):
        headerList = []
        for nowValue in valueList:
            tmpCellItem = ExcellCell(value=nowValue)
            tmpCellItem.setCellAsHeader()
            headerList.append(tmpCellItem)
        return headerList

    # 传入一个值的列表，将其转换成excell单元格列表
    def transformListToCellList(self, valueList, border="thin", fontAlignment="align", fontStyle="none",
                                fgColor="white"):
        cellList = []
        for nowValue in valueList:
            tmpCellItem = ExcellCell(value=nowValue, border=border, fontAlignment=fontAlignment, fontStyle=fontStyle,
                                     fgColor=fgColor)
            cellList.append(tmpCellItem)
        return cellList

    # 创建一个excell文件对象,并返回该文件对象
    def addFile(self, fileName="导出文件", fileSuffix="xlsx", sheetName=""):
        tmpExcellFile = ExcellFile(fileName=fileName, fileSuffix=fileSuffix, sheetName=sheetName)
        self.fileList.append(tmpExcellFile)
        return tmpExcellFile

    def getFinalFile(self):
        return self.fileList[len(self.fileList) - 1]

    # 写入一个空格单元格，防止上一列文本超出
    def writeExcellSpaceCell(self, ws, row, column):
        # 设置值
        ws.cell(row=row, column=column).value = " "
        return ws

    # 写入一个内容单元格
    def writeExcellCell(self, ws, row, column, cellItem):
        value = str(cellItem.value)
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        # 获得常用样式
        styleDic = self.getStyleDict()
        # 获得指定单元格
        aimCell = ws.cell(row=row, column=column)
        # 设置值
        aimCell.value = value
        # 设置边框
        styleObj = styleDic["borders"][cellItem.border]
        aimCell.border = styleObj
        # 设置居中
        aimCell.alignment = styleDic["fontAlignments"][cellItem.fontAlignment]

        # 设置超链接
        if cellItem.hyperLink:
            # 写入超链接
            aimCell.hyperlink = cellItem.hyperLink
            # 设置当前单元格字体类型为超链接类型
            aimCell.font = styleDic["fontStyles"]["hyper"]
        else:
            aimCell.font = styleDic["fontStyles"][cellItem.fontStyle]

        # 设置填充颜色
        fill = PatternFill("solid", fgColor=styleDic["colorCodes"][cellItem.fgColor])
        aimCell.fill = fill

        return ws

    # 删除指定路径的文件,传入一个绝对路径,返回一个布尔变量以及一个字符串变量，
    # 布尔变量为True表示是否删除成功,若为False则字符串变量中写入错误信息
    def deleteFile(self, filePath):
        deleteFlag = True
        reStr = ""
        if os.path.exists(filePath):
            try:
                os.remove(filePath)
            except Exception as ex:
                reStr = "删除失败，失败信息为：{0}".format(ex)
                deleteFlag = False
        else:
            reStr = "未找到指定路径的文件"
            deleteFlag = False
        return deleteFlag, reStr

    # 保存excell文件
    def saveExcell(self, wb, saveName):
        # 处理传入的文件名
        savePath = "{0}\\{1}".format(os.getcwd(), saveName)

        # 检测当前目录下是否有该文件，如果有则清除以前保存文件
        if os.path.exists(savePath):
            self.deleteFile(savePath)
        wb.save(savePath)
        return True

    # 设置指定列的列宽，colIndex从0开始
    def setExcellColWidth(self, ws, colIndex, colWidth):
        ws.column_dimensions[chr(ord("A") + colIndex)].width = colWidth
        return ws

    # 导出当前类中的excell文件，传入需导入的索引，索引从0开始，传入-1表示导出所有文件
    def exportExcell(self, fileIndex=-1):
        print("\n------开始导出数据至excell文件------")
        exportIndexs = []
        if fileIndex == -1 or fileIndex >= len(self.fileList):
            exportIndexs = range(0, len(self.fileList))
        else:
            exportIndexs.append(fileIndex)
        for tmpExportListIndex, exportIndex in enumerate(exportIndexs):
            nowFileItem = self.fileList[exportIndex]
            print(
                "【Excell文件】正在导出第{0}个文件（{0}/{1}）:{2}".format(tmpExportListIndex + 1, len(exportIndexs),
                                                                      nowFileItem.getSaveName()))
            try:
                # 创建一个excell文件对象
                nowWb = oxl.Workbook()
                # 遍历所有sheet
                nowSheets = nowFileItem.getSheets()
                for tmpSheetIndex, nowSheetItem in enumerate(nowSheets):
                    print(
                        "【Excell子表】正在导出第{0}个子表（{0}/{1}）:{2}".format(tmpSheetIndex + 1, len(nowSheets),
                                                                              nowSheetItem.getSheetName()))
                    try:
                        if tmpSheetIndex == 0:
                            nowWs = nowWb.active
                            nowWs.title = nowSheetItem.getSheetName()
                        else:
                            nowWs = nowWb.create_sheet(nowSheetItem.getSheetName(), tmpSheetIndex)
                        # 写入表头
                        nowHeaderList = nowSheetItem.getHeaderList()
                        print("【Excell数据】正在导出表头")
                        try:
                            for tmpCellIndex, nowCellItem in enumerate(nowHeaderList):
                                self.writeExcellCell(nowWs, 1, tmpCellIndex + 1, nowCellItem)
                        except Exception as ex:
                            logStr = "【Excell数据】导出表头失败，报错信息为：{0}".format(traceback.format_exc())
                            print(logStr)
                        # 写入数据
                        nowRows = nowSheetItem.getRows()
                        for tmpRowIndex, cellItemList in enumerate(nowRows):
                            print(
                                "\r【Excell数据】正在导出第{0}行数据（{0}/{1}）".format(tmpRowIndex + 1, len(nowRows)),
                                end="")
                            try:
                                for tmpCellIndex, nowCellItem in enumerate(cellItemList):
                                    self.writeExcellCell(nowWs, tmpRowIndex + 2, tmpCellIndex + 1, nowCellItem)

                                # 指定数量行保存一次
                                if tmpRowIndex != 0 and tmpRowIndex % self.saveCount == 0:
                                    self.saveExcell(nowWb, saveName=nowFileItem.getSaveName())
                                    nowWb = oxl.load_workbook(nowFileItem.getSaveName())
                                    nowWs = nowWb[nowWb.sheetnames[tmpSheetIndex]]
                            except Exception as ex:
                                logStr = "\n【Excell数据】导出第{0}行数据（{0}/{1}） 失败，报错信息为：{2}".format(
                                    tmpRowIndex + 1,
                                    len(nowRows),
                                    traceback.format_exc())
                                print(logStr)
                        print("，导出数据完成")
                        # 计算并设置列宽
                        for tmpColIndex in range(len(nowHeaderList)):
                            tmpWidthList = []
                            print(
                                "【Excell数据】正在自动设置第{0}列数据列宽（{0}/{1}）".format(tmpColIndex + 1,
                                                                                          len(nowHeaderList)),
                                end="")
                            try:
                                nowHeaderItem = nowHeaderList[tmpColIndex]
                                tmpWidthList.append(nowHeaderItem.getCellWidth())
                                nowColList = [row[tmpColIndex] for row in nowRows]
                                nowColWidthList = [cell.getCellWidth() for cell in nowColList]
                                tmpWidthList += nowColWidthList
                                tmpWidthList.sort(key=lambda width: int(width), reverse=True)
                                calcedWidth = int(tmpWidthList[0])
                                nowMaxCellWidth = self.styleDict["cellWidths"]["max"]
                                if calcedWidth > nowMaxCellWidth:
                                    calcedWidth = nowMaxCellWidth
                                self.setExcellColWidth(nowWs, tmpColIndex, calcedWidth)
                                print("，成功设置列宽为{0}".format(calcedWidth))
                            except Exception as ex:
                                logStr = "，设置失败，报错信息为：{0}".format(traceback.format_exc())
                                print(logStr)

                        # 设置冻结第一行
                        nowWs.freeze_panes = "B2"

                        print("【Excell子表】第{0}个子表（{0}/{1}）:{2} 导出完成".format(tmpSheetIndex + 1, len(nowSheets),
                                                                                     nowSheetItem.getSheetName()))
                    except Exception as ex:
                        logStr = "【Excell子表】导出第{0}个子表（{0}/{1}）:{2} 失败，报错信息为：{3}".format(
                            tmpSheetIndex + 1,
                            len(nowSheets),
                            nowSheetItem.getSheetName(),
                            traceback.format_exc())
                        print(logStr)
                # 保存未保存数据
                self.saveExcell(nowWb, saveName=nowFileItem.getSaveName())
                print("【Excell文件】第{0}个文件（{0}/{1}）:{2} 成功导出至当前文件夹".format(tmpExportListIndex + 1,
                                                                                         len(exportIndexs),
                                                                                         nowFileItem.getSaveName()))
            except Exception as ex:
                logStr = "【Excell文件】导出第{0}个文件（{0}/{1}）:{2} 失败，报错信息为：{3}".format(tmpExportListIndex + 1,
                                                                                               len(exportIndexs),
                                                                                               nowFileItem.getSaveName(),
                                                                                               traceback.format_exc())
                print(logStr)
        print("------完成导出数据至excell文件------\n")
